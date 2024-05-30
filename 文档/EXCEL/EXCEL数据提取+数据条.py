import openpyxl
import xlsxwriter
import os

# 打开原始Excel文件
input_file = "小盘价值barra2因子测试结论.xlsx"
wb = openpyxl.load_workbook(input_file, data_only=True)

# 创建新的输出文件名
base_name = os.path.splitext(input_file)[0]
output_file = f"{base_name}_数据提取+数据条.xlsx"

# 创建一个新的Excel文件
workbook = xlsxwriter.Workbook(output_file)

for sheetname in wb.sheetnames:
    worksheet = wb[sheetname]
    output_worksheet = workbook.add_worksheet(sheetname)

    # 获取T2:T6的数据
    data = []
    for row in range(2, 7):
        cell_value = worksheet[f'T{row}'].value
        data.append(cell_value)

    # 创建百分比格式
    percent_format = workbook.add_format({'num_format': '0.00%'})

    # 将数据写入新的工作表并设置为百分比格式
    for row, value in enumerate(data, start=1):
        output_worksheet.write(row, 0, value, percent_format)

    # 应用数据条条件格式
    data_bar = {
        'type': 'data_bar',
        'bar_solid': True,
        'bar_color': '#638EC6',  # 蓝色
        'min_type': 'num',        # 使用数字最小值
        'max_type': 'num',        # 使用数字最大值
        'min_value': min(data) if min(data) < 0 else 0,  # 如果最小值小于0，则使用实际最小值，否则使用0
        'max_value': max(data), 
        'bar_negative_color': '#FF6347'  # 红色
    }

    output_worksheet.conditional_format(1, 0, 5, 0, data_bar)

# 关闭工作簿
workbook.close()

print(f"转换完成，文件保存为 {output_file}")
