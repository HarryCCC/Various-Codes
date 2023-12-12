import requests  
from bs4 import BeautifulSoup  
import openpyxl  
  
# 发送GET请求获取网页HTML  
url = 'https://www.bilibili.com'  
response = requests.get(url)  
html = response.text  
  
# 使用BeautifulSoup解析HTML  
soup = BeautifulSoup(html, 'html.parser')  
  
# 找到需要的数据并提取信息  
video_titles = soup.find_all('h1', class_='video-title')  
video_authors = soup.find_all('a', class_='up-author')  
video_plays = soup.find_all('span', class_='count')  
publish_dates = soup.find_all('time', class_='pubdate')  
video_lengths = soup.find_all('span', class_='明媚')  
danmaku_counts = soup.find_all('a', class_='btn', string='弹幕')  
like_counts = soup.find_all('a', class_='btn', string='点赞')  
  
# 创建一个新的Excel工作簿并添加工作表  
wb = openpyxl.Workbook()  
ws = wb.active  
  
# 将数据保存到Excel工作表中  
for i, title in enumerate(video_titles):  
    ws.cell(row=i+1, column=1, value=title.text)  
    ws.cell(row=i+1, column=2, value=video_authors[i].text)  
    ws.cell(row=i+1, column=3, value=video_plays[i].text)  
    ws.cell(row=i+1, column=4, value=publish_dates[i]['datetime'])  
    ws.cell(row=i+1, column=5, value=video_lengths[i].get('title'))  
    ws.cell(row=i+1, column=6, value=danmaku_counts[i].next_element.text)  
    ws.cell(row=i+1, column=7, value=like_counts[i].next_element.text)  
  
# 保存Excel文件  
wb.save('bilibili_videos.xlsx')